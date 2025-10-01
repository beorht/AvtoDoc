from openpyxl import Workbook, load_workbook
# from pypandoc import convert_text
from google import genai
from time import sleep
from os import mkdir


# GEMINI_API_KEY
API_KEY = "AIzaSyBftd0C_yf4KyD0Nk2Vt9YXSO3aBhmwDRk"

# The client gets the API key from the environment variable `GEMINI_API_KEY`.
client = genai.Client(api_key=API_KEY)


# Создание файла MarkDown с лекционным материалом
def writeMD( filename: str, text: str ):

    filename += ".md"
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            f.write( text )
        
        print(f"::: Сгенерированный Markdown-отчет сохранен в файл '{filename}'")
    
    except IOError as e:
        print(f"Ошибка при сохранении файла: {e}")


def writeWord( filename: str, text: str ):
    filename += ".docx"

    convert_text(
        text,
        'docx',
        format="md",
        outputfile=filename,
        extra_args=['--standalone'] # --standalone создаст полный документ Word
    )

    print(f"::: Сгенерированный Word-отчет сохранен в файл '{filename}'")



# Текстовый промпт для отправки в нейронку
def sendPrompt( title: str, index: int ) -> str:
    
    prompt_ru = f"""
Напиши образовательный лекционный материал по теме `{title}` в формате Markdown по следующему шаблону:

#### Тема № {index + 1}: [вставь тему]
Основные разделы образовательного материала:
(Например в таком формате, от двух до пяти, в самом конце образовательного материала должны быть закрепляющие тему вопросы)
##### План:
1. ...
2. ...
3. ...
4. Рефлексивные вопросы 

Текст образовательного материала:
[сформулируй развернутый текст с пояснениями, примерами и выводами по теме]
    """

    prompt_uz = f"""
Ta’limiy lektsion materialni `{title}` mavzusi bo‘yicha quyidagi shablon asosida yozing, ajratuvchi chiziqlar siz:

#### Mavzusi № {index + 1}: {title}
Ta’limiy materialning asosiy bo‘limlari:
(Masalan, shunday formatda: ikkitadan beshtagacha, ta’limiy material oxirida esa mavzuni mustahkamlovchi savollar bo‘lishi kerak)
##### Reja:
1. ...
2. ...
3. ...
4. Refleksiv savollar


Ta’limiy material matni:
[mavzu bo‘yicha tushuntirishlar, misollar va xulosalar bilan kengaytirilgan matnni shakllantiring]
    """

    ### Ответ от нейронки
    return client.models.generate_content(
        model="gemini-2.5-flash", contents=prompt_ru
    )


# Функция чтения тем из текстового файла
def readThemeFromText( filename="data/theme.txt" ) -> list:

    themes = []

    with open( filename, "r", encoding='utf-8') as f:
        
        for line_number, line in enumerate(f, 1): # enumerate для получения номера строки
            # Каждая строка будет включать символ нового строки (\n) в конце,
            # если он присутствует в файле.
            # Для удаления пробельных символов (включая \n) можно использовать .strip()
            themes.append(f"{line_number}. {line.strip()}")

    print( "::: Темы из текстового файла прочтены!!!" )

    return themes


# Функция чтения тем из exel файла
def readThemeFromExel( filename="data/theme.xlsx" ) -> list:

    themes = []

    # ✅ Открытие существующего Excel
    wb2 = load_workbook( filename )
    ws2 = wb2.active

    # Чтение данных
    for row in ws2.iter_rows(values_only=True):
        
        if row[0] == None:
            continue
        
        themes.append(f"{row[0]}")

    print( "::: Темы из Exel файла прочтены!!!" )

    return themes


def main():
    
    while True:

        # Получаем название дисциплины
        subject_name = input( "\nВведите название предмета: " )

        if subject_name == "q":
            break

        mkdir( subject_name )  # Создаем папку под дисциплину 
        sleep( 1 )

        # themes = readThemeFromText()   # Получаем список тем с txt
        themes = readThemeFromExel()   # Получаем список тем с exel

        obshiy = ""

        for i in range( len( themes ) ):
            response = sendPrompt( themes[i], i )

            writeMD( f"{subject_name}/{i}-Leksiya", response.text )
            # writeWord( f"{subject_name}/{theme}", response.text )

            obshiy += "\n" + response.text

            sleep( 3 )

        writeMD( f"{subject_name}/Obshiy", obshiy )

if __name__ == '__main__':
    main()